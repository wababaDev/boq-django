import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError

from domain.catalog.models import Trade
from domain.boq.models import Project, BillItem

User = get_user_model()

SHEETS_TO_IMPORT = [
    "2 Bedroom Semi Detached - 140M2",
    "External Works ",
]


class Command(BaseCommand):
    help = "Import Wanji's sample 2-bed semi BoQ as a real test project, for validating the app against her actual work."

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str)
        parser.add_argument("--username", type=str, required=True)

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["filepath"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {options['filepath']}")

        try:
            user = User.objects.get(username=options["username"])
        except User.DoesNotExist:
            raise CommandError(f"No user named '{options['username']}'")

        trade_lookup = {t.name.strip().upper(): t for t in Trade.objects.all()}
        unsorted_trade, _ = Trade.objects.get_or_create(
            name="Unsorted (needs review)", defaults={"order": 999}
        )

        project = Project.objects.create(
            name="2 Bedroom Semi Detached - Sample",
            client_name="Sample Client",
            location="Sample Site",
            created_by=user,
        )

        total, unsorted_count = 0, 0
        for sheet_name in SHEETS_TO_IMPORT:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"Sheet not found: {sheet_name}"))
                continue
            count, unsorted = self._import_sheet(wb[sheet_name], project, trade_lookup, unsorted_trade)
            total += count
            unsorted_count += unsorted
            self.stdout.write(f"  {sheet_name}: {count} items ({unsorted} unsorted)")

        self.stdout.write(self.style.SUCCESS(
            f"Done. Project '{project.name}' (id={project.pk}) — {total} items, "
            f"{unsorted_count} landed in 'Unsorted' and need manual trade re-assignment."
        ))

    def _import_sheet(self, ws, project, trade_lookup, unsorted_trade):
        count, unsorted_count = 0, 0
        current_trade = None

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            item_code, description, unit, qty, rate = (c.value for c in row[:5])

            # Heading row: text in description, nothing in item_code
            if description and not item_code:
                heading = str(description).strip().upper()
                if heading in trade_lookup:
                    current_trade = trade_lookup[heading]
                continue

            # Real item row: has a code, description, unit, and numeric qty
            if item_code and description and unit and isinstance(qty, (int, float)):
                trade = current_trade or unsorted_trade
                if trade is unsorted_trade:
                    unsorted_count += 1
                BillItem.objects.create(
                    project=project,
                    trade=trade,
                    description=str(description).strip(),
                    unit=str(unit).strip(),
                    quantity=qty,
                    rate=rate or 0,
                )
                count += 1
        return count, unsorted_count