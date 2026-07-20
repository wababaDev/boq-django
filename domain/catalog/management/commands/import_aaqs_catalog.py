import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from domain.catalog.models import Trade, CatalogItem

SKIP_SHEETS = {"Cover page", "Final summary", "Preliminaries", "Contents", "Preface"}


class Command(BaseCommand):
    help = "Import AAQS model catalog from the reference Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str)
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete existing Trade/CatalogItem rows before importing.",
        )

    def handle(self, *args, **options):
        filepath = options["filepath"]
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {filepath}")

        if options["clear"]:
            self.stdout.write("Clearing existing catalog data...")
            CatalogItem.objects.all().delete()
            Trade.objects.all().delete()

        total_items = 0

        with transaction.atomic():
            for order, sheet_name in enumerate(wb.sheetnames):
                if sheet_name in SKIP_SHEETS:
                    continue

                ws = wb[sheet_name]
                trade, _ = Trade.objects.get_or_create(
                    name=sheet_name,
                    defaults={"order": order},
                )

                count = self._import_sheet_items(ws, trade)
                total_items += count
                self.stdout.write(f"  {sheet_name}: {count} items")

                if count == 0:
                    self.stdout.write(self.style.WARNING(
                        f"    (no items found — check this sheet manually)"
                    ))

        self.stdout.write(self.style.SUCCESS(
            f"Done. {total_items} catalog items imported across "
            f"{Trade.objects.count()} trades."
        ))

    def _import_sheet_items(self, ws, trade):
        """A real item row has an integer item number in column A
        and a unit in column C. Everything else (section subheadings,
        preamble paragraphs, blank rows) has item_no = None."""
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            item_no = row[0].value
            description = row[1].value
            unit = row[2].value

            if not isinstance(item_no, (int, float)):
                continue
            if not description or not str(description).strip():
                continue

            CatalogItem.objects.create(
                trade=trade,
                item_code=str(int(item_no)),
                description_template=str(description).strip(),
                unit=str(unit).strip() if unit else "",
            )
            count += 1
        return count